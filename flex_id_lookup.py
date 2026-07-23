#!/usr/bin/env python3
"""
Read an Excel table (header on row 2, columns: Date, Symbol, Order Qty, Side),
look each row up in a kdb+ `target` table via pykx, extract FIX tag 9603 from the
`fixmsg` column, and write a new xlsx with an added `FLEX ID` column.

The original workbook is opened with openpyxl and only ONE new column is added,
so every existing cell (values, styles, number formats, formulas) is preserved
byte-for-byte in the output file.

Rows that resolve to exactly one kdb match get their FLEX ID.
Rows with zero or several matches are left blank in the xlsx AND listed in a
conflicts CSV so they can be resolved by hand.

Usage
-----
    python flex_id_lookup.py orders.xlsx --host myhost --port 5000

    python flex_id_lookup.py orders.xlsx --host myhost --port 5000 \
        --sheet Sheet1 --out orders_with_flex.xlsx --conflicts conflicts.csv
"""

import argparse
import csv
import sys
from copy import copy

import openpyxl
import pykx as kx

# --------------------------------------------------------------------------- #
# CONFIG — adjust these to match your kdb schema if the defaults don't fit.
# --------------------------------------------------------------------------- #

# Name of the kdb table to query.
TARGET_TABLE = "target"

# The FIX tag to extract from the `fixmsg` column.
FIX_TAG = "9603"

# How the `side` column is stored in kdb, and how to translate the Excel value.
#   - SIDE_MAP: optional {excel_value: kdb_value} translation. Leave {} to pass
#     the Excel value through unchanged. Example for FIX codes:
#         SIDE_MAP = {"Buy": "1", "Sell": "2"}
#   - SIDE_Q_TYPE: the q type to send the side as — "symbol" or "char".
#     Use "char"  if kdb `side` holds single chars like "1"/"2" or "B"/"S".
#     Use "symbol" if kdb `side` holds symbols like `Buy/`Sell or `1/`2.
SIDE_MAP: dict[str, str] = {
    "BUY": "buy",
    "SELL": "sell",
    "SSH": "sellshort",
}
SIDE_Q_TYPE = "symbol"

# Header text used for the column names on row 2 of the Excel file.
COL_DATE = "Date"
COL_SYMBOL = "Symbol"
COL_QTY = "Order Qty"
COL_SIDE = "Side"

# Row (1-indexed) that holds the column headers.
HEADER_ROW = 2

# Name of the new column added to the output workbook.
NEW_COL = "FLEX ID"

# --------------------------------------------------------------------------- #
# kdb / q side
# --------------------------------------------------------------------------- #

# getTag[tag;msg]: split the fix message on ";" (fields look like ";TAG=VALUE;")
# and pull the value of `tag`. Returns "" if the tag is absent.
_GETTAG_Q = r'''
getTag:{[tag;msg]
  f:";" vs msg;
  m:f where f like (tag,"=*");
  $[count m; (1+count tag)_ first m; ""] }
'''

# The constrained lookup, run DIRECTLY per row (no stored q function). Each key
# value is rendered as a q literal and dropped straight into the query, so its
# type is exactly what we write and nothing gets re-typed on the way through a
# lambda's arguments. FLEX is extracted server-side via getTag so we never ship
# whole fix messages back over the wire.
QUERY_TEMPLATE = (
    "select date,sym,size,side,basket,"
    'flex:getTag["' + FIX_TAG + '";] each fixmsg '
    "from " + TARGET_TABLE + " "
    'where date={date}, sym={sym}, size={size}, side={side}, '
    'basket like "*ARROWP*"'
)


def connect(host: str, port: int) -> kx.SyncQConnection:
    conn = kx.SyncQConnection(host=host, port=port)
    conn(_GETTAG_Q)  # getTag is a pure string helper — no query, no key types
    return conn


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #

def q_date(value) -> str:
    """The Excel date is already in q's 'YYYY.MM.DD' form, so pass it straight
    through as the q date literal."""
    return str(value).strip()


def q_sym(value) -> str:
    """Render a value as a q symbol literal, e.g. `AAPL."""
    return "`" + str(value).strip()


def q_qty(value) -> str:
    """Render the order quantity as a numeric q literal.

    Excel often stores quantities as text with thousands separators like
    '1,000'; strip the commas before parsing so it becomes a clean number.
    Whole values render as integers (q longs), fractional values as floats."""
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    num = float(value)
    return str(int(num) if num.is_integer() else num)


def q_side(value) -> str:
    """Translate + render the side value as a q literal (symbol or char)."""
    raw = str(value).strip()
    v = SIDE_MAP.get(raw.upper(), raw) if SIDE_MAP else raw
    if SIDE_Q_TYPE == "char":
        return '"' + v + '"'  # q char literal, e.g. "1"/"2" or "B"/"S"
    return "`" + v            # q symbol literal, e.g. `buy


def qstr(value) -> str:
    """Normalise a q string / symbol / bytes result to a Python str."""
    if isinstance(value, bytes):
        return value.decode("utf-8", "replace")
    return "" if value is None else str(value)


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #

def main() -> int:
    ap = argparse.ArgumentParser(description="Add FLEX ID (FIX tag 9603) from kdb to an Excel table.")
    ap.add_argument("excel", help="input .xlsx file")
    ap.add_argument("--host", required=True, help="kdb host")
    ap.add_argument("--port", required=True, type=int, help="kdb port")
    ap.add_argument("--sheet", default=None, help="worksheet name (default: active sheet)")
    ap.add_argument("--out", default=None, help="output .xlsx (default: <input>_with_flex.xlsx)")
    ap.add_argument("--conflicts", default=None, help="conflicts CSV (default: <input>_conflicts.csv)")
    args = ap.parse_args()

    out_path = args.out or args.excel.rsplit(".", 1)[0] + "_with_flex.xlsx"
    conflicts_path = args.conflicts or args.excel.rsplit(".", 1)[0] + "_conflicts.csv"

    # Load preserving everything (formulas kept: data_only=False is the default).
    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet] if args.sheet else wb.active

    # Locate the columns we need by their header text on HEADER_ROW.
    headers = {}
    for cell in ws[HEADER_ROW]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column
    for required in (COL_DATE, COL_SYMBOL, COL_QTY, COL_SIDE):
        if required not in headers:
            print(f"ERROR: column {required!r} not found on row {HEADER_ROW}. "
                  f"Found: {list(headers)}", file=sys.stderr)
            return 1

    c_date = headers[COL_DATE]
    c_sym = headers[COL_SYMBOL]
    c_qty = headers[COL_QTY]
    c_side = headers[COL_SIDE]

    # New column goes just after the last used column.
    new_col = ws.max_column + 1
    hdr_cell = ws.cell(row=HEADER_ROW, column=new_col, value=NEW_COL)
    # Blend the new header's look with the existing header to its left (only
    # touches the brand-new cell — never the original data).
    left = ws.cell(row=HEADER_ROW, column=new_col - 1)
    if left.has_style:
        hdr_cell.font = copy(left.font)
        hdr_cell.fill = copy(left.fill)
        hdr_cell.border = copy(left.border)
        hdr_cell.alignment = copy(left.alignment)
        hdr_cell.number_format = left.number_format

    conn = connect(args.host, args.port)

    conflicts = []
    n_matched = n_conflict = n_skipped = 0

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        d = ws.cell(row=row, column=c_date).value
        s = ws.cell(row=row, column=c_sym).value
        q = ws.cell(row=row, column=c_qty).value
        sd = ws.cell(row=row, column=c_side).value

        # Skip fully blank rows.
        if d is None and s is None and q is None and sd is None:
            continue
        # Skip rows missing any key field (can't match on partial keys).
        if d is None or s is None or q is None or sd is None:
            n_skipped += 1
            conflicts.append([row, d, s, q, sd, "incomplete-row", ""])
            continue

        try:
            query = QUERY_TEMPLATE.format(
                date=q_date(d),
                sym=q_sym(s),
                size=q_qty(q),
                side=q_side(sd),
            )
            res = conn(query).pd()
        except Exception as exc:  # noqa: BLE001 - report and keep going
            n_skipped += 1
            conflicts.append([row, d, s, q, sd, f"query-error: {exc}", ""])
            continue

        flex_values = [qstr(v) for v in res["flex"].tolist()] if "flex" in res else []
        n = len(res)

        if n == 1:
            ws.cell(row=row, column=new_col, value=flex_values[0])
            n_matched += 1
        else:
            # 0 or >1 matches -> leave FLEX ID blank, record for manual review.
            n_conflict += 1
            reason = "no-match" if n == 0 else f"multiple({n})"
            conflicts.append([row, d, s, q, sd, reason, " | ".join(flex_values)])

    wb.save(out_path)
    print(f"Wrote {out_path}  (matched={n_matched}, conflicts={n_conflict}, skipped={n_skipped})")

    if conflicts:
        with open(conflicts_path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["excel_row", COL_DATE, COL_SYMBOL, COL_QTY, COL_SIDE,
                        "reason", "candidate_flex_ids"])
            w.writerows(conflicts)
        print(f"Wrote {conflicts_path}  ({len(conflicts)} rows to resolve)")
    else:
        print("No conflicts — every row resolved to exactly one match.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
